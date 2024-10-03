import pandas as df
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

class ExcelReader:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.data = None

    def read_equipment_data(self) -> List[Dict[str, str]]:
        """Lee los datos de los equipos desde la hoja de cálculo."""
        try:
            # Lee el archivo Excel directamente en un DataFrame de pandas
            self.data = df.read_excel(self.file_path)
            
            # Convierte el DataFrame a una lista de diccionarios
            equipment_data = self.data.to_dict('records')
            
            # Filtra las filas que tienen todos los valores
            equipment_data = [eq for eq in equipment_data if all(eq.values())]
            
            # Elimina columnas con nombres vacíos
            columns = [col for col in list(equipment_data[0].keys()) if col.strip()]
            
            return equipment_data
        except Exception as e:
            print(f"Error al leer el archivo Excel: {e}")
            return []

    def print_equipment_data(self, data: List[Dict[str, str]]):
        """Imprime los datos de los equipos de forma legible."""
        for equipment in data:
            print("\nEquipo:")
            for key, value in equipment.items():
                print(f"  {key}: {value}")

    def save_router_info(self, info: Dict[str, str]):
        if info:
            try:
                # Intentar cargar el archivo existente, si no existe, crear uno nuevo
                if os.path.exists(self.file_path):
                    wb = load_workbook(self.file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Device Info"
                    # Añadir encabezados si es un archivo nuevo
                    headers = ["Nombre", "Modelo", "Número de Serie", "Versión Firmware", "IP Router", "Conexion"]
                    ws.append(headers)

                # Añadir la nueva información
                new_row = [
                    info['nombre'],
                    info['modelo'],
                    info['numero_serie'],
                    info['version_firmware'],
                    info['ip_router'],
                    info['connection_status']
                ]
                ws.append(new_row)

                # Ajustar el ancho de las columnas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Guardar el archivo
                wb.save(self.file_path)
                print(f"Información guardada en {self.file_path}")
                return f"Información guardada en {self.file_path}"
            except Exception as e:
                error_message = f"Error al guardar la información: {e}"
                print(error_message)
                return error_message

if __name__ == "__main__":
    # Ejemplo de uso
    reader = ExcelReader("equipos.xlsx")
    data = reader.read_equipment_data()
    reader.print_equipment_data(data)