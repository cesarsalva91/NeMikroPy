from librouteros import connect
from librouteros.exceptions import TrapError
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

class RouterManager:
    def __init__(self, router_ip, username, password, port):
        self.router_ip = router_ip
        self.username = username
        self.password = password
        self.port = port
        self.api = None
        self.connection_status = None  # Nuevo atributo para almacenar el estado de la conexión

    def connect_to_router(self):
        try:
            # Conexión al router MikroTik
            self.api = connect(username=self.username, password=self.password, host=self.router_ip, port=self.port)
            self.connection_status = "Conectado"
            return True
        except TrapError as e:
            error_message = f"Error en MikroTik API (TrapError) al conectar con {self.router_ip}:{self.port}: {e}"
            print(error_message)
            self.connection_status = error_message
        except Exception as e:
            error_message = f"Error inesperado al conectar con {self.router_ip}:{self.port}: {e}"
            print(error_message)
            self.connection_status = error_message
        return False

    def get_router_info(self):
        if not self.api:
            print("No hay conexión al router. Llame a connect_to_router() primero.")
            return None

        try:
            # Obtener información del sistema
            identity = list(self.api.path('system', 'identity'))
            system_resource = list(self.api.path('system', 'resource'))
            
            # Imprimir la información recibida para depuración
            print("Respuesta de 'system identity':", identity)
            print("Respuesta de 'system resource':", system_resource)

            # Extraer la información
            nombre = identity[0].get('name', 'Desconocido') if identity else 'Desconocido'
            modelo = system_resource[0].get('board-name', 'Desconocido') if system_resource else 'Desconocido'
            numero_serie = system_resource[0].get('serial-number', 'Desconocido') if system_resource else 'Desconocido'
            version_firmware = system_resource[0].get('version', 'Desconocido') if system_resource else 'Desconocido'
            
            return {
                'nombre': nombre,
                'modelo': modelo,
                'numero_serie': numero_serie,
                'version_firmware': version_firmware,
                'ip_router': self.router_ip,
                'connection_status': self.connection_status
            }
        except Exception as e:
            error_message = f"Error al obtener información del router: {e}"
            print(error_message)
            self.connection_status = error_message
            return None

    def save_router_info(self, info):
        if info:
            filename = "Network_Devices_List.xlsx"
            try:
                # Intentar cargar el archivo existente, si no existe, crear uno nuevo
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Device Info"
                    # Añadir encabezados si es un archivo nuevo
                    headers = ["Nombre", "Modelo", "Número de Serie", "Versión Firmware", "IP Router","Conexion"]
                    ws.append(headers)

                # Añadir la nueva información
                new_row = [
                    info['nombre'],
                    info['modelo'],
                    info['numero_serie'],
                    info['version_firmware'],
                    info['ip_router'],
                    info['connection_status']
                ]
                ws.append(new_row)

                # Ajustar el ancho de las columnas
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # Guardar el archivo
                wb.save(filename)
                print(f"Información guardada en {filename}")
                self.connection_status = f"Información guardada en {filename}"
            except Exception as e:
                error_message = f"Error al guardar la información: {e}"
                print(error_message)
                self.connection_status = error_message

    def run(self):
        if self.connect_to_router():
            info = self.get_router_info()
            if info:
                self.save_router_info(info)
        return self.connection_status  # Devuelve el estado de la conexión

# Uso de la clase
if __name__ == "__main__":
    # Parámetros de conexión
    router_ip = "192.168.240.134"  # Cambia esto por la IP de tu equipo MikroTik
    username = "admin"         # Cambia esto por tu usuario
    password = "admin"              # Cambia esto por tu contraseña
    port = 4444                # Puerto API por defecto, cambia si es necesario

    # Crear instancia y ejecutar
    router_manager = RouterManager(router_ip, username, password, port)
    connection_status = router_manager.run()
    print(f"Estado de la conexión: {connection_status}")